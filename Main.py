

import pandas as pd
from openpyxl import load_workbook
import os
from util import *
# -

pathname = "G:\My Drive\CV-2021\Applications 2021\BRAINOMIX-Algorithm Researcher\BRAINOMIX challenge\Images/"
print(test)
pathname_results = r"G:\My Drive\CV-2021\Applications 2021\BRAINOMIX-Algorithm Researcher\BRAINOMIX challenge\Results.xlsx"
foldername = "vol_09/"
filename = "vol_09"
directory = os.path.isdir(pathname + foldername)
if not directory:
    os.mkdir(pathname + foldername)


ext = '.nii.gz'

# load image
nii_img = nib.load(pathname + filename + ext)

nii_data = nii_img.get_fdata()

'get dimensions'
dim = (nii_img.header.get_data_shape())
'get num of slices to loop'
num_slices = dim[2]
'aray to store results'

lung_area_arr = np.zeros((num_slices))

lung_area_arr_ml = np.zeros((num_slices))
lung_area_ml_total = np.zeros(1)
vessel_vol = np.zeros((num_slices))

vessel_vol_ml = np.zeros((num_slices))
vessel_vol_ml_total = np.zeros(1)

lung_vessel_ratio = np.zeros((num_slices))
lung_vessel_ratio_total=np.zeros(1)

combi_mask = np.zeros([dim[0], dim[1], dim[2]])

'get voxel size to calculate volume'
vox = (nii_img.header.get_zooms())
'loop over the volume'
for i in range(num_slices):

    image = nii_data[:, :, i]

    'plot figure'
    #plt.figure(figsize=(100, 100))
    #plt.style.use('grayscale')
    #plt.imshow(image)
    #plt.show()

    'binarize'

    upper_threshold = -300
    lower_threshold = -1000

    'call function to binarize the image into lung and not lung tissue'
    image_bin = binarize(image, upper_threshold, lower_threshold)

    'display the binarized image'
    #plt.figure()
    #plt.imshow(image_bin.T)
    #plt.show()

    ''''as vessels are present in the lung tissue simple thresholding based on binary values  would
 not provide accurate lung volume. Contour of the lung is required.
 create contour of the lung based on the binary image'''''

    contours = create_contour(image_bin, image)

    ''''
  separate the lung contour from the rest of the contours including the region outside the body
 '''

    lung_contour = lungs_cont(contours)

    #fig, ax1 = plt.subplots()
    #ax1.imshow(image)

    #for contour in lung_contour:
        #ax1.plot(contour[:, 1], contour[:, 0], linewidth=2)

    #plt.show()

    lung_mask = create_mask(image, lung_contour)
    #plt.figure()
    #plt.imshow(lung_mask)
    #plt.show
    combi_mask[:,:,i] = lung_mask
    out_mask_name = (pathname + foldername + filename + "_mask" + str(i))

    save_nifty(lung_mask, out_mask_name, nii_img.affine)
    area_lung = lung_area(lung_mask, vox)
    lung_area_arr[i] = area_lung

    lung_area_arr_ml[i] = (area_lung / 1000)
    lung_area_ml_total += lung_area_arr_ml[i]

    main_vessels, vessel_vol[i] = sep_main_vessels(lung_mask.T, image, vox)


    vessel_vol_ml[i]= vessel_vol[i]/1000
    vessel_vol_ml_total += vessel_vol_ml[i]

    lung_vessel_ratio[i]= (vessel_vol_ml[i] / lung_area_arr_ml[i])*100
    lung_vessel_ratio_total += lung_vessel_ratio[i]

combi_name = (pathname+ foldername + filename + "_ mask")
save_nifty(combi_mask, combi_name, nii_img.affine)

results_df = pd.DataFrame(
    {'lung slice volume': lung_area_arr,
     'lung slice volume ml': lung_area_arr_ml,
     'vessel_vol mm3': vessel_vol,
     'vessel_vol ml': vessel_vol_ml,
     'lung_vessel_ratio (%)': lung_vessel_ratio})
result_total_df = pd.DataFrame({'Volume number':filename,
                                'Total lung volume (ml)':lung_area_ml_total,
                                'vessel volume total (ml)':vessel_vol_ml_total,
                                'lung vessel ratio (%)':lung_vessel_ratio_total})
results_book = load_workbook(pathname_results)

writer = pd.ExcelWriter(pathname_results, engine='openpyxl',mode='a',if_sheet_exists='overlay')

writer.book = results_book
results_df.to_excel(writer, sheet_name= filename)

#if 'Combined Results' in results_book.sheetnames:
 #   ws=results_book['Combined Results1']
 #   newRowLocation = ws.max_row + 1

    # write to the cell you want, specifying row and column, and value :-)
 #   ws.cell(column=1, row=newRowLocation, value=result_total_df[1])
    #result_total_df.to_excel(writer, sheet_name='Combined Results1')
#else:
 #    results_book.create_sheet('Combined Results1')
  #   result_total_df.to_excel(writer, sheet_name='Combined Results1')



#results_book = load_workbook(pathname_results)
#writer = pd.ExcelWriter(pathname_results, engine='openpyxl',mode='a',if_sheet_exists='overlay')

#writer.book = results_book
#result_total_df.to_excel(writer, sheet_name= 'Combined Results')


#results_book = load_workbook(pathname_results)
#writer2 = pd.ExcelWriter(pathname_results, engine='openpyxl', mode='a')
#writer2.book = results_book






#append_df_to_excel:(pathname_results, results_total_df, sheet_name='Combined Results', index=False, startrow=2)

#writer.save()
writer.close()
#df_complete_read = pd.read_excel(pathname_results, sheet_name='Combined Results')
#df_complete_appended=df_complete_read.append(result_total_df)
#results_book = load_workbook(pathname_results)
#del results_book ['Combined Results']
#writer = pd.ExcelWriter(pathname_results, engine='openpyxl')
#writer.book = results_book
#df_complete_appended.to_excel(writer, sheet_name='Combined Results')
#writer.close()
print(results_df)
print(result_total_df)


